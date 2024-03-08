import openpyxl 
import pandas as pd  
import pandas as pd
from genderize import Genderize 

path = "Msc_CA_Oct-2023_Ledger1.xlsx" 
wb_obj = openpyxl.load_workbook(path)
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
 
sheet_obj = wb_obj.active 
 

# def predict_gender(name):
#     try:
#         gender = Genderize().get([name])[0]['gender']
#         return gender
#     except Exception as e:
#         print(f"Error predicting gender for {name}: {e}")
#         return None

# excel_filename = 'output_data.xlsx'
# df = pd.read_excel(excel_filename)
# names_to_predict = df["name"].tolist()
# mid_names = [word.split()[1] if len(word.split()) > 1 else word for word in names_to_predict]
# print("Last names:", mid_names)
name_gender = [
"female","male","female","female","male","male","male","male","male","male","male","male","male","male","female","male","male","male","male","female","male","male","male","male","male","female","male","male","male","male","female","male","male","female","male","male","male","female","male","male","male","female","female","female","male","male","male","female","male","female","female","male","female","male","male","female","male","female"]
# for name in mid_names:
#     gender = predict_gender(name)
#     name_gender.append(gender)
#     print(f"{name}: {gender}")
print(len(name_gender))
  

subjects_Code1 = sheet_obj.cell(row=2,column=5) 
subjects_Code2 = sheet_obj.cell(row=2,column=6)
subjects_Code3 = sheet_obj.cell(row=2,column=7)
subjects_Code4 = sheet_obj.cell(row=11,column=5)  
subjects_Code5 = sheet_obj.cell(row=11,column=6)
subjects_Code6 = sheet_obj.cell(row=11,column=7)
subjects_Code7 = sheet_obj.cell(row=20,column=5)
subjects_Code8 = sheet_obj.cell(row=20,column=6)
subjects_Code9 = sheet_obj.cell(row=20,column=7)
subject_code1 = subjects_Code1.value
subject_code2 = subjects_Code2.value
subject_code3 = subjects_Code3.value
subject_code4 = subjects_Code4.value
subject_code5 = subjects_Code5.value
subject_code6 = subjects_Code6.value
subject_code7 = subjects_Code7.value
subject_code8 = subjects_Code8.value
subject_code9 = subjects_Code9.value
new_row = ["name", 
           'gender',
           "seat_no", 
            subject_code1 ,
            subject_code2,
            subject_code3,
            subject_code4,
            subject_code5,
            subject_code6,
            subject_code7,
            subject_code8, 
            subject_code9,
           "SGPA",
           "percentage"
           
           
           ]
new_sheet.append(new_row)


sub_count = 2
count = 5
gradecount = 8
row = 8
incount = 0
gcount=0
while(row <= sheet_obj.max_row-2):
     
        while count <= sheet_obj.max_column:
            stu_info = sheet_obj.cell(row=1, column=count)
            Marks_info = sheet_obj.cell(row=2, column=count+3)  
            obtained_marks = sheet_obj.cell(row=11, column=count+4)  

            sub1 = sheet_obj.cell(row=8,column=5+incount).value
            sub2 = sheet_obj.cell(row=8,column=6+incount).value
            sub3 = sheet_obj.cell(row=8,column=7+incount).value
            sub4 = sheet_obj.cell(row=17,column=5+incount).value
            sub5 = sheet_obj.cell(row=17,column=6+incount).value
            sub6 = sheet_obj.cell(row=17,column=7+incount).value
            sub7 = sheet_obj.cell(row=26,column=5+incount).value
            sub8 = sheet_obj.cell(row=26,column=6+incount).value
            sub9 = sheet_obj.cell(row=26,column=7+incount).value
              
            name = stu_info.value.split("\n")[0]
            seat_no = stu_info.value.split("\n")[3]   
            SGPA = Marks_info.value.split("  ")[1].split(" ")[1] 
            result =   ( int(obtained_marks.value.split(":")[1]) / 500)
            percentage = f"{result:.2%}"
            new_sheet.append([name,name_gender[gcount],seat_no,sub1,sub2,sub3,sub4,sub5,sub6,sub7,sub8,sub9,SGPA,percentage])
            count += 5
            sub_count +=1
            row = row+9
            incount += 5
            gcount += 1
         

new_file_path = "output_data.xlsx"
new_wb.save(new_file_path)

df = pd.read_excel("output_data.xlsx")
toppers_wb = openpyxl.Workbook()
five_toppers = toppers_wb.active
top_five_students = df.sort_values(by='percentage', ascending=False).head(5)
five_toppers.append(['name',"percentage"])

  

for index, row in top_five_students.iterrows():

    five_toppers.append([row["name"] ,row["percentage"]])
    print(f"{row['name']} - {row['percentage']}%")  

 
toppers_file_path = "toppers.xlsx"
toppers_wb.save(toppers_file_path)
 