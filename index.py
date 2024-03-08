import openpyxl 
import pandas as pd  

df = pd.read_excel("./output_data.xlsx")
print(df)

main_wb = openpyxl.Workbook()

toppers = main_wb.create_sheet("Toppers")
 
top_five = df.sort_values(by='percentage', ascending=False).head(5)
toppers.append(['name','percentage'])

for index, row in top_five.iterrows():
    toppers.append([row["name"],row["percentage"]])
    print(f"{row['name']} - {row['percentage']}%")




SubjectWise = main_wb.create_sheet("SubjectWise")
  
  
grade_counts = {
    'O': 'First class with Distinction (above 70%)',
    'A+': 'First class (60-69%)',
    'A': 'Higher Second class (55-59%)',
    'B+': 'Second class (50-54%)',
    'B': 'Pass class (40-49%)',
    'F': 'Fail'
}
 
subjects = ['22-CACCTP-7', '23-CA-CBOTP-3A','23-392','23-CACCTP-8','23-CA-CBOPP-3A','23-394','23-CACCTP-9','23-CA-CCPP-3','23-395']
SubjectWise.append([''] + subjects)
 
 
    
for grade in grade_counts:
    grade_counts_subject = []
    for subject in subjects:
        grades = df[subject]
        count = len(grades[grades == grade])
        grade_counts_subject.append(count)
    SubjectWise.append([grade_counts[grade]] + grade_counts_subject)
 


overall = main_wb.create_sheet("OverAll")



 
         
        
    
    




genderWise = main_wb.create_sheet("GenderWise")

male = df[df['gender'] =='male']

main_wb.save("result.xlsx")